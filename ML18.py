from __future__ import unicode_literals, print_function, division
import pickle
import re
from sklearn.metrics import confusion_matrix
from sklearn.linear_model import LogisticRegression
from sklearn.linear_model import Perceptron
from sklearn.linear_model  import PassiveAggressiveClassifier
from sklearn.neural_network import MLPClassifier
from sklearn.svm import SVC
from sklearn.feature_extraction.text import CountVectorizer
import openpyxl
import numpy as np
import tqdm
import warnings
import random
import multiprocessing as mp

warnings.filterwarnings("ignore")
datafrom = openpyxl.load_workbook('NSCLCPD-1.xlsx')
train = datafrom.get_sheet_by_name("Training")

AccuracyMLP = []
AccuracyLogistic = []
x_pointless = []
data = []
for t in range(2,2252):
    a=train.cell(row = t, column=12).value
    b=a[0:a.rfind("ictated")-1]
    if (train.cell(row = t,column = 3).value in ("POD","SD")):
        c = 0
    else:
        c = 1
    data.append((b,c))
random.shuffle(data)
avgmlp = []
trainData = [data[j] for j in range(0,2050)]
testData = [data[u] for u in range(2051,2250)]
trainText, trainY = [d[0] for d in trainData], [d[1] for d in trainData]
testText, testY = [d[0] for d in testData], [d[1] for d in testData]
min_df = 1
max_features = 15000
countVec = CountVectorizer(ngram_range=(1,3), min_df = min_df, max_features = max_features)
trainX = countVec.fit_transform(trainText)
testX = countVec.transform(testText)
mlp = MLPClassifier(hidden_layer_sizes=(100,), alpha = .0001, batch_size='auto', learning_rate='constant', learning_rate_init=0.001, power_t=0.5,max_iter=2000,shuffle=True, random_state=None, tol=0.0001, momentum=0.9)
mlp.fit(trainX, trainY)
a = mlp.predict(testX)
falpos = 0
falneg = 0
truesc = 0
for i in tqdm.tqdm(range(len(testY))):
    b = mlp.predict(testX[i])
    c = int(b)
    d = testY[i]
    if (c != d):
        if (c == 0):
            print((i,"false positive"))
            falpos += 1
        else:
            print((i,"false negative"))
            falneg +=1
    else:
        truesc += 1
falposper = (falpos/(len(testY)))
falnegper = (falneg/(len(testY)))
truescper = (truesc/(len(testY)))
print((falpos,falneg,truesc))
print((falposper,falnegper,truescper))
mlp_score = mlp.score(testX, testY)
print(mlp_score)
